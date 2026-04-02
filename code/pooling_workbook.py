"""
Build Illumina-style pooling Excel workbooks from templates + read targets.
"""

from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Template layout (1-based Excel rows) — matches 2025-12-03_pooling_WS.xlsx & Pool1_2
# Yellow summary: label column B (2), value column C (3). Row numbers differ by template.
META_VALUE_COL = 3
META_ROWS_ATAC = {
    "pool_name": 11,
    "n_samples": 12,
    "final_conc_lib": 13,
    "pool_nM": 14,
    "pool_vol": 15,
}
META_ROWS_RNA = {
    "pool_name": 10,
    "n_samples": 11,
    "final_conc_lib": 12,
    "pool_nM": 13,
    "pool_vol": 14,
}
# RNA / small RNA: first library row is Excel row 17 (header row 16).
DATA_START_ROW_RNA = 17
# ATAC Pool1_2: header row 17, first library row 18.
DATA_START_ROW_ATAC = 18


def sample_sort_key(sid: str) -> tuple:
    parts = sid.split("_")
    out: list = []
    for x in parts:
        out.append(int(x) if x.isdigit() else x)
    return tuple(out)


def sorted_sample_ids_from_atac(eff_df: pd.DataFrame) -> list[str]:
    ids = eff_df["sample_id"].astype(str).unique().tolist()
    return sorted(ids, key=sample_sort_key)


def parse_atac_template_meta(
    excel_path: Path, sorted_ids: list[str]
) -> dict[str, dict]:
    df = pd.read_excel(excel_path, sheet_name="Pool1_2", header=None)
    meta: dict[str, dict] = {}
    pat = re.compile(r"ATAC-(\d+)\s*$", re.I)
    for i in range(17, min(df.shape[0], 60)):
        name = df.iloc[i, 1]
        if pd.isna(name):
            continue
        s = str(name).strip()
        if s.lower() == "sum":
            break
        m = pat.search(s)
        if not m:
            continue
        idx = int(m.group(1))
        if idx < 1 or idx > len(sorted_ids):
            continue
        sid = sorted_ids[idx - 1]
        meta[sid] = {
            "display_name": s,
            "conc_ng_ul": float(df.iloc[i, 2]),
            "size_bp": float(df.iloc[i, 3]),
            "molarity_nM": float(df.iloc[i, 4]),
            "tube_well": None,
        }
    return meta


def parse_rna_template_meta(excel_path: Path, sorted_ids: list[str]) -> dict[str, dict]:
    df = pd.read_excel(excel_path, sheet_name="RNA", header=None)
    meta: dict[str, dict] = {}
    pat = re.compile(r"Nov25-(\d+)\s*$")
    for i in range(16, min(df.shape[0], 60)):
        name = df.iloc[i, 1]
        if pd.isna(name):
            continue
        s = str(name).strip()
        m = pat.search(s)
        if not m:
            continue
        idx = int(m.group(1))
        if idx < 1 or idx > len(sorted_ids):
            continue
        sid = sorted_ids[idx - 1]
        tw = df.iloc[i, 6]
        meta[sid] = {
            "display_name": s,
            "conc_ng_ul": float(df.iloc[i, 2]),
            "size_bp": float(df.iloc[i, 3]),
            "molarity_nM": float(df.iloc[i, 4]),
            "tube_well": None if pd.isna(tw) else str(tw),
        }
    return meta


def parse_srna_template_meta(excel_path: Path) -> dict[str, dict]:
    df = pd.read_excel(excel_path, sheet_name="small RNA", header=None)
    meta: dict[str, dict] = {}
    pat = re.compile(r"Nov25-(\d+)-sRNA-(ZI|FR)_(\d+)_(\d+)")
    for i in range(16, min(df.shape[0], 60)):
        name = df.iloc[i, 1]
        if pd.isna(name):
            continue
        s = str(name).strip()
        m = pat.search(s)
        if not m:
            continue
        sid = f"{m.group(1)}_{m.group(2)}_{m.group(3)}_{m.group(4)}"
        tw = df.iloc[i, 6]
        meta[sid] = {
            "display_name": s,
            "conc_ng_ul": float(df.iloc[i, 2]),
            "size_bp": float(df.iloc[i, 3]),
            "molarity_nM": float(df.iloc[i, 4]),
            "tube_well": None if pd.isna(tw) else str(tw),
        }
    return meta


def _col_letter(col_idx: int) -> str:
    return get_column_letter(col_idx)


def _meta_value_abs(row: int) -> str:
    """Absolute reference to summary value column C (e.g. $C$14)."""
    return f"${_col_letter(META_VALUE_COL)}${row}"


def molarity_correction_factor(
    sample_id: str,
    demux_sample_counts: dict[str, int],
    *,
    warn_label: str = "",
) -> float:
    """
    (observed share of assignable reads) / (1/n) for the prior multiplex pool,
    from demux sample_counts (prompt.md rule 9).
    """
    if not demux_sample_counts:
        return 1.0
    total = float(sum(demux_sample_counts.values()))
    n = len(demux_sample_counts)
    if total <= 0 or n <= 0:
        return 1.0
    r = float(demux_sample_counts.get(sample_id, 0))
    if sample_id not in demux_sample_counts and warn_label:
        print(
            f"WARNING: {warn_label}: sample {sample_id} missing from demux sample_counts; "
            "molarity correction factor = 1",
            file=sys.stderr,
        )
        return 1.0
    observed = r / total
    expected = 1.0 / n
    f = observed / expected if expected > 0 else 1.0
    if f <= 0:
        if warn_label:
            print(
                f"WARNING: {warn_label}: sample {sample_id} has no assignable reads in demux; "
                "molarity correction factor = 1",
                file=sys.stderr,
            )
        return 1.0
    return f


def _final_conc_per_lib_formula(meta_rows: dict) -> str:
    """Pool total nM / number of libraries (yellow block). Safe when n_samples is 0."""
    c_n = _meta_value_abs(meta_rows["n_samples"])
    c_nm = _meta_value_abs(meta_rows["pool_nM"])
    return f'=IF({c_n}>0,{c_nm}/{c_n},"n/a")'


def _strip_data_rows(ws, data_start_row: int) -> None:
    """Remove all rows from data_start_row through current max_row (old samples + sum rows)."""
    n = ws.max_row - data_start_row + 1
    if n > 0:
        ws.delete_rows(data_start_row, n)


def _find_sum_row(ws, data_start_row: int) -> int | None:
    """First row at/after data_start_row whose cell is exactly 'sum' (case-insensitive)."""
    for r in range(data_start_row, min(ws.max_row, data_start_row + 200) + 1):
        for c in range(1, 16):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "sum":
                return r
    return None


def _strip_sample_rows_keep_footer(ws, data_start_row: int) -> bool:
    """
    Delete old library rows only, keep footer (sum / add … / Total).
    Returns True if a 'sum' row was found; False if the sheet had no footer (full strip).
    """
    fr = _find_sum_row(ws, data_start_row)
    if fr is None:
        _strip_data_rows(ws, data_start_row)
        return False
    n_del = fr - data_start_row
    if n_del > 0:
        ws.delete_rows(data_start_row, n_del)
    return True


def _insert_sample_row_block(ws, data_start_row: int, n: int) -> None:
    if n > 0:
        ws.insert_rows(data_start_row, n)


def _truncate_rows_after(ws, last_kept_row: int) -> None:
    extra = ws.max_row - last_kept_row
    if extra > 0:
        ws.delete_rows(last_kept_row + 1, extra)


def _write_atac_footer(ws, ds: int, de: int, sum_row: int) -> None:
    """Pool1_2-style footer: sum, add Tris (= pool vol − pooled lib µL, same pattern as RNA add H2O)."""
    c_pool_atac = _meta_value_abs(META_ROWS_ATAC["pool_vol"])
    # Sample block: J = volume for pool, M = mass (after molarity correction columns).
    ws.cell(row=sum_row, column=7, value="sum")
    ws.cell(row=sum_row, column=8, value=f"=SUM(J{ds}:J{de})")
    for c in range(9, 13):
        ws.cell(row=sum_row, column=c).value = None
    ws.cell(row=sum_row, column=13, value=f"=SUM(M{ds}:M{de})")
    ws.cell(row=sum_row + 1, column=7, value="add Tris")
    for c in range(8, 14):
        ws.cell(row=sum_row + 1, column=c).value = None
    ws.cell(row=sum_row + 1, column=8, value=f"={c_pool_atac}-H{sum_row}")
    ws.cell(
        row=sum_row + 1,
        column=10,
        value="if <0 pool concentration required",
    )
    ws.cell(row=sum_row + 2, column=7, value="TOTAL")
    for c in range(8, 14):
        ws.cell(row=sum_row + 2, column=c).value = None
    ws.cell(row=sum_row + 2, column=8, value=f"=SUM(H{sum_row}:H{sum_row + 1})")
    _truncate_rows_after(ws, sum_row + 2)


def _write_rna_srna_footer(ws, ds: int, de: int, sum_row: int) -> None:
    """ATAC-aligned columns: sum, add H20, TOTAL, Concentration(ng)."""
    c_pool = _meta_value_abs(META_ROWS_RNA["pool_vol"])
    for r in range(sum_row, sum_row + 4):
        for c in range(7, 14):
            ws.cell(row=r, column=c).value = None
    ws.cell(row=sum_row, column=7, value="sum")
    ws.cell(row=sum_row, column=8, value=f"=SUM(J{ds}:J{de})")
    ws.cell(row=sum_row, column=13, value=f"=SUM(M{ds}:M{de})")
    ws.cell(row=sum_row + 1, column=7, value="add H20")
    ws.cell(row=sum_row + 1, column=8, value=f"={c_pool}-H{sum_row}")
    ws.cell(row=sum_row + 1, column=10, value="if <0 pool concentration required")
    ws.cell(row=sum_row + 2, column=7, value="TOTAL")
    ws.cell(row=sum_row + 2, column=8, value=f"=SUM(H{sum_row}:H{sum_row + 1})")
    ws.cell(row=sum_row + 3, column=7, value="Concentration(ng)")
    ws.cell(row=sum_row + 3, column=8, value=f"=SUM(M{ds}:M{de})/H{sum_row + 2}")
    _truncate_rows_after(ws, sum_row + 3)


# Sample block columns A–M (after insert of molarity correction pair).
POOL_SAMPLE_HEADER_LABELS: dict[int, str] = {
    1: "proportion_reads_within_pool",
    2: "sample name",
    3: "Conc undiluted (ng/μl)",
    4: "Size (bp)",
    5: "Molarity (nM)",
    6: "Molarity correction factor",
    7: "Molarity corrected based on previous sequencing (nM)",
    8: " Undiluted volume for pool (μl)",
    9: "Dilution factor",
    10: " Volume for pool (μl)",
    11: "Volume aliquot undiluted (uL)",
    12: "Volume of water added to dilute (uL)",
    13: "Mass of DNA  (ng)",
}


def copy_worksheet_shallow(src_ws, dst_wb, title: str):
    """Copy cell values and basic styles into a new sheet (good enough for templates)."""
    dst = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            c = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    c.font = copy(cell.font)
                    c.border = copy(cell.border)
                    c.fill = copy(cell.fill)
                    c.number_format = copy(cell.number_format)
                    c.protection = copy(cell.protection)
                    c.alignment = copy(cell.alignment)
                except Exception:
                    pass
    for col, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst.column_dimensions[col].width = dim.width
    for row, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst.row_dimensions[row].height = dim.height
    return dst


def fill_atac_pool_sheet(
    ws,
    pool_title: str,
    sample_ids: list[str],
    proportion_reads_within_pool: list[float],
    meta: dict[str, dict],
    V_pool_ul: float,
    pool_nM: float,
    demux_sample_counts: dict[str, int],
    overrides: dict[str, dict] | None = None,
    *,
    data_start_row: int = DATA_START_ROW_ATAC,
):
    """A–G: read share, IDs, conc/size, molarity, demux correction, corrected nM; H+ use column G for nM."""
    overrides = overrides or {}
    n = len(sample_ids)
    mr = META_ROWS_ATAC
    pool_vol_row = mr["pool_vol"]
    final_conc_row = mr["final_conc_lib"]

    had_footer = _strip_sample_rows_keep_footer(ws, data_start_row)

    ws.cell(row=mr["pool_name"], column=META_VALUE_COL, value=pool_title)
    ws.cell(row=mr["n_samples"], column=META_VALUE_COL, value=n)
    ws.cell(
        row=final_conc_row,
        column=META_VALUE_COL,
        value=_final_conc_per_lib_formula(mr),
    )
    ws.cell(row=mr["pool_nM"], column=META_VALUE_COL, value=pool_nM)
    ws.cell(row=pool_vol_row, column=META_VALUE_COL, value=V_pool_ul)

    if n == 0:
        return

    ws.insert_cols(6, 2)
    _insert_sample_row_block(ws, data_start_row, n)

    concs: list[float] = []
    sizes: list[float] = []
    for sid in sample_ids:
        ovr = overrides.get(sid, {})
        m = dict(meta.get(sid, {}))
        if "conc_ng_ul" in ovr:
            m["conc_ng_ul"] = ovr["conc_ng_ul"]
        if "size_bp" in ovr:
            m["size_bp"] = ovr["size_bp"]
        concs.append(float(m.get("conc_ng_ul", 0)))
        sizes.append(float(m.get("size_bp", 0)))

    dr_end = data_start_row + n - 1
    c_pool = _meta_value_abs(pool_vol_row)
    c_nm = _meta_value_abs(mr["pool_nM"])

    hdr_row = data_start_row - 1
    for col, label in POOL_SAMPLE_HEADER_LABELS.items():
        ws.cell(row=hdr_row, column=col, value=label)

    for k, sid in enumerate(sample_ids):
        row = data_start_row + k
        fac = molarity_correction_factor(
            sid, demux_sample_counts, warn_label=pool_title
        )
        ws.cell(row=row, column=1, value=float(proportion_reads_within_pool[k]))
        ws.cell(row=row, column=2, value=sid)
        ws.cell(row=row, column=3, value=concs[k])
        ws.cell(row=row, column=4, value=sizes[k])
        ws.cell(
            row=row,
            column=5,
            value=f"=C{row}*1000000/(660*D{row})",
        )
        ws.cell(row=row, column=6, value=float(fac))
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        ws.cell(
            row=row,
            column=8,
            value=f"={c_pool}*({c_nm}*A{row})/G{row}",
        )
        ws.cell(
            row=row,
            column=9,
            value=f"=IF(H{row}>=1,1,MAX(2,ROUNDUP(1/H{row},0)))",
        )
        ws.cell(row=row, column=10, value=f"=H{row}*I{row}")
        ws.cell(row=row, column=11, value=f'=IF(I{row}>1,1.5,"")')
        ws.cell(row=row, column=12, value=f'=IF(I{row}>1,K{row}*(I{row}-1),"")')
        ws.cell(row=row, column=13, value=f"=C{row}*H{row}")

    sum_row = data_start_row + n
    if had_footer:
        _write_atac_footer(ws, data_start_row, dr_end, sum_row)
    else:
        _truncate_rows_after(ws, dr_end)


def fill_rna_like_pool_sheet(
    ws,
    pool_title: str,
    sample_ids: list[str],
    proportion_reads_within_pool: list[float],
    meta: dict[str, dict],
    V_pool_ul: float,
    pool_nM: float,
    demux_sample_counts: dict[str, int],
    *,
    data_start_row: int = DATA_START_ROW_RNA,
):
    """RNA / small RNA: same column layout as ATAC (A–M); undiluted volume uses corrected nM (G)."""
    n = len(sample_ids)
    mr = META_ROWS_RNA
    pool_vol_row = mr["pool_vol"]
    final_conc_row = mr["final_conc_lib"]

    had_footer = _strip_sample_rows_keep_footer(ws, data_start_row)

    extra_right = ws.max_column - 10
    if extra_right > 0:
        ws.delete_cols(11, extra_right)

    ws.delete_cols(7, 1)
    ws.insert_cols(9, 2)
    ws.insert_cols(6, 2)

    hdr_row = data_start_row - 1
    for col, label in POOL_SAMPLE_HEADER_LABELS.items():
        ws.cell(row=hdr_row, column=col, value=label)

    for r in (mr["n_samples"], mr["final_conc_lib"]):
        for c in range(5, 14):
            ws.cell(row=r, column=c).value = None

    ws.cell(row=mr["pool_name"], column=META_VALUE_COL, value=pool_title)
    ws.cell(row=mr["n_samples"], column=META_VALUE_COL, value=n)
    ws.cell(
        row=final_conc_row,
        column=META_VALUE_COL,
        value=_final_conc_per_lib_formula(mr),
    )
    ws.cell(row=mr["pool_nM"], column=META_VALUE_COL, value=pool_nM)
    ws.cell(row=pool_vol_row, column=META_VALUE_COL, value=V_pool_ul)

    if n == 0:
        return

    _insert_sample_row_block(ws, data_start_row, n)

    concs: list[float] = []
    sizes: list[float] = []
    for sid in sample_ids:
        m = meta.get(sid, {})
        concs.append(float(m.get("conc_ng_ul", 0)))
        sizes.append(float(m.get("size_bp", 0)))

    dr_end = data_start_row + n - 1
    c_pool = _meta_value_abs(pool_vol_row)
    c_nm = _meta_value_abs(mr["pool_nM"])

    for k, sid in enumerate(sample_ids):
        row = data_start_row + k
        fac = molarity_correction_factor(
            sid, demux_sample_counts, warn_label=pool_title
        )
        ws.cell(row=row, column=1, value=float(proportion_reads_within_pool[k]))
        ws.cell(row=row, column=2, value=sid)
        ws.cell(row=row, column=3, value=concs[k])
        ws.cell(row=row, column=4, value=sizes[k])
        ws.cell(row=row, column=5, value=f"=C{row}*1000000/(660*D{row})")
        ws.cell(row=row, column=6, value=float(fac))
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        ws.cell(row=row, column=8, value=f"={c_pool}*({c_nm}*A{row})/G{row}")
        ws.cell(
            row=row,
            column=9,
            value=f"=IF(H{row}>=1,1,MAX(2,ROUNDUP(1/H{row},0)))",
        )
        ws.cell(row=row, column=10, value=f"=H{row}*I{row}")
        ws.cell(row=row, column=11, value=f'=IF(I{row}>1,1.5,"")')
        ws.cell(row=row, column=12, value=f'=IF(I{row}>1,K{row}*(I{row}-1),"")')
        ws.cell(row=row, column=13, value=f"=C{row}*H{row}")

    sum_row = data_start_row + n
    if had_footer:
        _write_rna_srna_footer(ws, data_start_row, dr_end, sum_row)
    else:
        _truncate_rows_after(ws, dr_end)


def read_template_meta_rows(ws, meta_rows: dict) -> tuple[float, float]:
    def _num(cell):
        v = cell.value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    pool_nM = _num(ws.cell(row=meta_rows["pool_nM"], column=META_VALUE_COL))
    V_pool = _num(ws.cell(row=meta_rows["pool_vol"], column=META_VALUE_COL))
    return (pool_nM if pool_nM is not None else 30.0), (
        V_pool if V_pool is not None else 30.0
    )


def _filter_meta(
    sample_ids: list[str],
    proportion_reads_within_pool: list[float],
    meta: dict[str, dict],
    label: str,
) -> tuple[list[str], list[float]]:
    out_s: list[str] = []
    out_p: list[float] = []
    for sid, p in zip(sample_ids, proportion_reads_within_pool):
        if sid not in meta:
            print(
                f"WARNING: {label}: sample {sid} not in pooling template; skipped in xlsx.",
                file=sys.stderr,
            )
            continue
        out_s.append(sid)
        out_p.append(float(p))
    return out_s, out_p


def build_pooling_workbook(
    out_path: Path,
    atac_template: Path,
    rna_srna_template: Path,
    sorted_atac_ids: list[str],
    pools: dict[str, dict],
    demux_sample_counts_by_omic: dict[str, dict[str, int]],
) -> None:
    """
    pools keys: ATAC-seq_1, ATAC-seq_2, RNA-seq, sRNA-seq
    demux_sample_counts_by_omic keys: ATAC-seq, RNA-seq, sRNA-seq (full prior multiplex pool).
    each pool value: {
        "sample_ids": [...],
        "proportion_reads_within_pool": [...],
        "overrides": {sid: {conc_ng_ul, size_bp}} (optional)
    }
    """
    wb_at = load_workbook(atac_template, data_only=False)
    wb_rs = load_workbook(rna_srna_template, data_only=False)

    nM_atac, V_atac = read_template_meta_rows(wb_at["Pool1_2"], META_ROWS_ATAC)
    nM_rna, V_rna = read_template_meta_rows(wb_rs["RNA"], META_ROWS_RNA)

    meta_at = parse_atac_template_meta(atac_template, sorted_atac_ids)
    meta_rna = parse_rna_template_meta(rna_srna_template, sorted_atac_ids)
    meta_srna = parse_srna_template_meta(rna_srna_template)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # ATAC sheets (from ATAC template)
    for title, key in (
        ("ATAC-seq_1", "ATAC-seq_1"),
        ("ATAC-seq_2", "ATAC-seq_2"),
    ):
        spec = pools[key]
        s_f, p_f = _filter_meta(
            spec["sample_ids"],
            spec["proportion_reads_within_pool"],
            meta_at,
            f"ATAC pool {title}",
        )
        ws = copy_worksheet_shallow(wb_at["Pool1_2"], wb_out, title)
        fill_atac_pool_sheet(
            ws,
            f"{title} add-on pool",
            s_f,
            p_f,
            meta_at,
            V_atac,
            nM_atac,
            demux_sample_counts_by_omic["ATAC-seq"],
            overrides=spec.get("overrides"),
        )

    # RNA
    spec = pools["RNA-seq"]
    s_f, p_f = _filter_meta(
        spec["sample_ids"],
        spec["proportion_reads_within_pool"],
        meta_rna,
        "RNA-seq pool",
    )
    ws = copy_worksheet_shallow(wb_rs["RNA"], wb_out, "RNA-seq")
    fill_rna_like_pool_sheet(
        ws,
        "RNA-seq add-on pool",
        s_f,
        p_f,
        meta_rna,
        V_rna,
        nM_rna,
        demux_sample_counts_by_omic["RNA-seq"],
    )

    spec = pools["sRNA-seq"]
    s_f, p_f = _filter_meta(
        spec["sample_ids"],
        spec["proportion_reads_within_pool"],
        meta_srna,
        "sRNA-seq pool",
    )
    ws = copy_worksheet_shallow(wb_rs["small RNA"], wb_out, "sRNA-seq")
    fill_rna_like_pool_sheet(
        ws,
        "sRNA-seq add-on pool",
        s_f,
        p_f,
        meta_srna,
        V_rna,
        nM_rna,
        demux_sample_counts_by_omic["sRNA-seq"],
    )

    wb_out.save(out_path)
